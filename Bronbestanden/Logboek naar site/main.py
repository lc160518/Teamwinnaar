from openpyxl import load_workbook

data_file = "C:/Users/eigen/OneDrive - Quadraam/PWS/logboek-python.xlsx"

# Load the entire workbook.
wb = load_workbook(data_file, data_only=True)
ws = wb["Blad1"]
datacollum = ws["F"]
all_rows = list(ws.rows)

def listall():
    # List all the sheets in the file.
    print("Found the following worksheets:")
    for sheetname in wb.sheetnames:
        print(sheetname)


def listrows():
    # Load one worksheet.
    ws = wb["Julian"]
    all_rows = list(ws.rows)

    print(f"Found {len(all_rows)} rows of data.")

    print("\nFirst rows of data:")
    for row in all_rows[:5]:
        print(row)

def viewdata():
    for i in range(len(datacollum)):
        print(datacollum[i].value)
        

